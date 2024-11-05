from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from app01 import models
from app01.utils.form import StudentForm
from app01.utils.pagination import Pagination


# 学生列表
def stu_list(request):
    form = StudentForm()
    data_dict = {}
    search_data = request.GET.get('q', '')
    if search_data:
        data_dict['username__contains'] = search_data
    queryset = models.Account.objects.filter(**data_dict)
    page_obj = Pagination(request, queryset)
    context = {
        'form': form,
        'queryset': page_obj.page_queryset,  # 分完页的数据
        'page_string': page_obj.html()  # 页码
    }
    return render(request, 'stu_list.html', context)


# 上传
def stu_multi(request):
    # 1.获取用户上传的文件对象
    file_obj = request.FILES.get('exc')

    # 2.对象传递给openpyxl，由它读取文件的内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        id_number = row[0].value
        name = row[1].value
        pwd = row[2].value
        exists = models.Account.objects.filter(id_number=id_number).exists()
        if not exists:
            models.Account.objects.create(id_number=id_number, username=name, password=pwd)
    return redirect('/student/list/')


# 添加
@csrf_exempt
def stu_add(request):
    form = StudentForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})

    return JsonResponse({'status': False, 'errors': form.errors})


# 编辑
def stu_edit(request):
    row_dict = models.Account.objects.filter(id=request.GET.get('uid')).values('username', 'id_number',
                                                                               'password', 'auth').first()
    context = {
        'status': True,
        'data': row_dict
    }
    return JsonResponse(context)


# 编辑保存
@csrf_exempt
def stu_edit_save(request):
    row_obj = models.Account.objects.filter(id=request.GET.get('edit_id')).first()
    form = StudentForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})

    return JsonResponse({'status': False, 'errors': form.errors})


# 删除
def stu_delete(request):
    models.Account.objects.filter(id=request.GET.get('uid')).delete()
    return JsonResponse({'status': True})